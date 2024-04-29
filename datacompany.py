from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook

file_path = r'D:\stockcode\code2\data.xlsx'
sheet_name = 'Dữ liệu'

# Load workbook
wb = load_workbook(file_path)

# Check if the "BCTC" sheet exists and create it if not
if "BCTC" not in wb.sheetnames:
    sheet_data = wb.create_sheet("BCTC")
else:
    sheet_data = wb["BCTC"]
    # Clear the existing data in the sheet
    sheet_data.delete_rows(sheet_data.min_row, sheet_data.max_row)
    sheet_data.delete_cols(sheet_data.min_column, sheet_data.max_column)

# Get data from the "Dữ liệu" sheet
sheet = wb[sheet_name]

# Get a list of existing symbols
existing_symbols = [value.value for value in sheet['A'][1:]]

# Set header for columns
sheet_data.cell(row=1, column=1, value="Mã")
sheet_data.cell(row=1, column=2, value="Tên")

# Create a list of years to fetch data for
start_year = 2021
end_year = 2023
custom_years = [str(year) for year in range(start_year, end_year + 1)]

# Set header for year columns in reversed order
for col, year in enumerate(reversed(custom_years), start=3):
    sheet_data.cell(row=1, column=col, value=year)

row = 2

# Create a dictionary to store data
data_dict = {}

for symbol in existing_symbols:
    for i, year in enumerate(custom_years):
        url = f'https://s.cafef.vn/Ajax/HoSoCongTy.aspx?symbol={symbol}&Type=2&PageIndex={i}&PageSize=4'
        response = requests.get(url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, "html.parser")
            data_rows = []
            # Loop through both index ranges to get data
            for index in range(9):
                data_rows.extend(soup.select(f"tr[id^=rptNhomChiTieu_rptData_0_TrData_{index}]"))
            for index in range(5):
                data_rows.extend(soup.select(f"tr[id^=rptNhomChiTieu_rptData_1_TrData_{index}]"))

            for tr_id in data_rows:
                first_td = tr_id.select_one("td:nth-of-type(1)")
                value_td = tr_id.select_one("td:nth-of-type(5)")
                if first_td and value_td:
                    name = first_td.text.strip()
                    value = value_td.text.strip()
                    if (symbol, name) not in data_dict:
                        data_dict[(symbol, name)] = {year: value}
                    else:
                        data_dict[(symbol, name)][year] = value

# Write data to the sheet
for (symbol, name), year_data in data_dict.items():
    sheet_data.cell(row=row, column=1, value=symbol)
    sheet_data.cell(row=row, column=2, value=name)
    for col, year in enumerate(reversed(custom_years), start=3):
        sheet_data.cell(row=row, column=col, value=year_data.get(year, ""))
    row += 1

# Save the workbook
wb.save(file_path)
