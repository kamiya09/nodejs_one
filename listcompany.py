import requests
import json
import re
from bs4 import BeautifulSoup
import openpyxl
import pandas as pd

def get_financial_data(symbol, normal):
    url2 = f'{normal}/bao-cao-tai-chinh/{symbol}/bsheet/2023/0/0/0/bao-cao-tai-chinh-.chn'
    url3 = f'{normal}/Ajax/CongTy/ThongTinChung.aspx?sym={symbol}'

    response2 = requests.get(url2)
    response3 = requests.get(url3)

    if response2.status_code == 200 and response3.status_code == 200:
        soup = BeautifulSoup(response2.content, 'html.parser')
        div_content = soup.find('div', {'id': 'div_cf_BoxContent'})
        if div_content:
            text = div_content.text.strip()
            match = re.search(r'\((.*?)\)', text)
            if match:
                text_in_parentheses = match.group(1)
                api_content = BeautifulSoup(response3.content, 'html.parser')
                charter_capital = api_content.find('span', {'id': 'ucThongTinChung1_lblTitle_Nhomnganh'}).next_sibling.strip()
                registered_capital = api_content.find('span', {'id': 'ucThongTinChung1_lblTitle_VonDieule'}).next_sibling.strip()
                registered_capital = re.sub(r'\s*đồng\s*', '', registered_capital)
                listed_stock_volume = api_content.find('span', {'id': 'ucThongTinChung1_Label1'}).next_sibling.strip()
                listed_stock_volume = re.sub(r'\s*cp\s*', '', listed_stock_volume)
                circulating_stock_volume = api_content.find('span', {'id': 'ucThongTinChung1_Label2'}).next_sibling.strip()
                circulating_stock_volume = re.sub(r'\s*cp\s*', '', circulating_stock_volume)
                return (text_in_parentheses, charter_capital, registered_capital, listed_stock_volume, circulating_stock_volume)
            else:
                return (None, None, None, None, None)
        else:
            return (None, None, None, None, None)
    else:
        return (None, None, None, None, None)

def main():
    many = "1654"
    normal = 'https://s.cafef.vn'
    url1 = f'{normal}/ajax/pagenew/databusiness/congtyniemyet.ashx?centerid=0&skip=0&take={many}&major=0'
    response = requests.get(url1)

    if response.status_code == 200:
        json_data = json.loads(response.content)
        data = json_data['Data']

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Dữ liệu"
        sheet['A1'] = 'Symbol'
        sheet['B1'] = 'Company Name'
        sheet['C1'] = 'Price'
        sheet['D1'] = 'Stock Exchange'
        sheet['E1'] = 'Category Name'
        sheet['F1'] = 'Charter Capital (Đồng)'
        sheet['G1'] = 'Listed Stock Volume (CP)'
        sheet['H1'] = 'Circulating Stock Volume (CP)'

        history_sheet = workbook.create_sheet("LSGD")
        history_sheet.append(['Symbol', 'Ngày', 'Giá điều chỉnh', 'Giá đóng cửa', 'Thay đổi', 'Khối lượng khớp lệnh', 'Giá trị khớp lệnh', 'Khớp lệnh thoả thuận', 'Giá trị thoả thuận', 'Giá mở cửa', 'Giá cao nhất', 'Giá thấp nhất'])

        data_rows = []  # Tạo một danh sách để lưu trữ dữ liệu cho từng hàng

        for item in data:
            symbol = item['Symbol']
            financial_data = get_financial_data(symbol, normal)

            if financial_data[0] is not None:
                data_rows.append({
                    'Symbol': symbol,
                    'Company Name': item['CompanyName'],
                    'Price': item['Price'],
                    'Stock Exchange': financial_data[0],
                    'Category Name': financial_data[1],
                    'Charter Capital (Đồng)': financial_data[2],
                    'Listed Stock Volume (CP)': financial_data[3],
                    'Circulating Stock Volume (CP)': financial_data[4]
                })

            else:
                print(f"Không thể lấy dữ liệu cho mã {symbol}")

            url4 = f'{normal}/Ajax/PageNew/DataHistory/PriceHistory.ashx?Symbol={symbol}&StartDate=&EndDate=&PageIndex=0&PageSize=100'
            response4 = requests.get(url4)
            if response4.status_code == 200:
                history_data = response4.json()['Data']['Data']
                if history_data:
                    for history in history_data:
                        history_sheet.append([
                            symbol,
                            history['Ngay'],
                            history['GiaDieuChinh'],
                            history['GiaDongCua'],
                            history['ThayDoi'],
                            history['KhoiLuongKhopLenh'],
                            history['GiaTriKhopLenh'],
                            history['KLThoaThuan'],
                            history['GtThoaThuan'],
                            history['GiaMoCua'],
                            history['GiaCaoNhat'],
                            history['GiaThapNhat']
                        ])

                else:
                    print(f"Không thể lấy dữ liệu lịch sử giao dịch cho mã {symbol}")

        # Chuyển danh sách thành DataFrame
        df = pd.DataFrame(data_rows)

        # Sắp xếp DataFrame theo cột "Category Name"
        df_sorted = df.sort_values(by='Category Name')

        # Lặp lại các hàng đã sắp xếp và in vào tệp Excel
        for index, row_data in df_sorted.iterrows():
            for col_index, value in enumerate(row_data):
                sheet.cell(row=index + 2, column=col_index + 1, value=value)

        file_path = 'D:\\stockcode\\code2\\data.xlsx'
        workbook.save(file_path)

        print("Done:", file_path)
    else:
        print(f"Yêu cầu thất bại cho url1 với mã lỗi: {response.status_code}")

if __name__ == "__main__":
    main()
