import pandas as pd
from openpyxl import load_workbook
import requests 
import json
# Đường dẫn tới tệp Excel và tên sheet
file_path = r'D:\stockcode\code2\data.xlsx'
sheet_name = 'Dữ liệu'

# Đọc dữ liệu từ cột "Symbol" trong sheet "Dữ liệu"
wb = load_workbook(file_path)
sheet = wb[sheet_name]
existing_symbols = [value for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True) for value in row]

def get_data_api():
    url = 'https://s.cafef.vn/ajax/mobile/smart/ajaxbandothitruong.ashx?type=1&category=0&centerId=0'
    response = requests.get(url)
    if response.status_code == 200:
        json_data = json.loads(response.content)
        data = json_data['Data']
        return data
    else:
        return None

# Lấy danh sách Symbol từ API
def get_api_symbols(data):
    api_symbols = [item['Symbol'] for item in data]
    return api_symbols

# Lấy thông tin của các công ty có Symbol trùng nhau giữa sheet "Dữ liệu" và API
def get_matching_companies(data, existing_symbols):
    api_symbols = get_api_symbols(data)
    matching_symbols = set(existing_symbols) & set(api_symbols)
    matching_companies = [item for item in data if item['Symbol'] in matching_symbols]
    return matching_companies

# Sử dụng các hàm
data = get_data_api()
if data:
    matching_companies = get_matching_companies(data, existing_symbols)
    if matching_companies:
        # Tạo DataFrame từ danh sách các công ty
        df = pd.DataFrame(matching_companies)

        # Xóa sheet "Data" cũ nếu tồn tại
        if "Data" in wb.sheetnames:
            wb.remove(wb["Data"])

        # Tạo sheet "Data" mới
        sheet_data = wb.create_sheet("Data")
        sheet_data['A1'] ='Symbol'
        sheet_data['B1'] ='Result'
        sheet_data['C1'] ='Color'
        sheet_data['D1'] ='Price'
        sheet_data['E1'] ='ChangePercent'
        sheet_data['F1'] ='Change'
        sheet_data['G1'] ='Name'
        sheet_data['H1'] ='TotalVolume'
        sheet_data['I1'] ='TotalValue'
        sheet_data['J1'] ='MarketCap'
       
        # Ghi dữ liệu vào sheet
        for row in df.itertuples(index=True, name='Pandas'):
            sheet_data.append([getattr(row, col) for col in df.columns])

        # Lưu file Excel
        wb.save(file_path)
        print(f"Dữ liệu đã được ghi vào file '{file_path}' trong sheet 'Data' mới.")
    else:
        print("Không có công ty nào có mã trùng nhau giữa sheet 'Dữ liệu' và API")
else:
    print("Không thể lấy dữ liệu từ API")